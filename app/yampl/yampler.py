import functools
import pandas as pd
from typing import List, Union, Callable, Any
from pathlib import Path
import yaml
import openpyxl


class TestIterator:
    """
    A decorator to test functions with multiple input variations without modifying the original function.
    
    Usage:
        @TestIterator("major_string", ["string1", "string2", ...])
        @TestIterator("major_string", "path/to/file.yaml")
        @TestIterator("major_string", "path/to/file.xlsx", excel_column="major_strings")
        def my_function(minor_str="default"):
            ...
    """
    
    def __init__(self, param_name: str, test_values: Union[List[str], str], 
                 yaml_key: str = None, excel_column: str = None, excel_sheet: str = None):
        """
        Initialize the decorator.
        
        Args:
            param_name: Name of the parameter to iterate (e.g., "major_string")
            test_values: Either a list of values, or path to YAML/Excel file
            yaml_key: If YAML file, the key to extract values from (optional, defaults to param_name)
            excel_column: If Excel file, the column name to read values from
            excel_sheet: If Excel file, the sheet name (defaults to first sheet)
        """
        self.param_name = param_name
        self.test_values_input = test_values
        self.yaml_key = yaml_key or param_name
        self.excel_column = excel_column
        self.excel_sheet = excel_sheet
        self.test_values = self._load_test_values()
        
    def _load_test_values(self) -> List[str]:
        """Load test values from list, YAML, or Excel file."""
        # If it's already a list, return it
        if isinstance(self.test_values_input, list):
            return self.test_values_input
        
        # Otherwise, treat it as a file path
        file_path = Path(self.test_values_input)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Handle YAML files
        if file_path.suffix in ['.yaml', '.yml']:
            with open(file_path, 'r') as f:
                data = yaml.safe_load(f)
                
            # Extract values based on yaml_key
            if isinstance(data, dict):
                if self.yaml_key in data:
                    values = data[self.yaml_key]
                    return values if isinstance(values, list) else [values]
                else:
                    raise KeyError(f"Key '{self.yaml_key}' not found in YAML file")
            elif isinstance(data, list):
                return data
            else:
                return [str(data)]
        
        # Handle Excel files
        elif file_path.suffix in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Select sheet
            if self.excel_sheet:
                ws = wb[self.excel_sheet]
            else:
                ws = wb.active
            
            # Find column index
            if not self.excel_column:
                raise ValueError("excel_column parameter is required for Excel files")
            
            # Get header row to find column
            headers = [cell.value for cell in ws[1]]
            if self.excel_column not in headers:
                raise ValueError(f"Column '{self.excel_column}' not found in Excel file")
            
            col_idx = headers.index(self.excel_column) + 1
            
            # Extract values from column (skip header)
            values = []
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_value = row[0].value
                if cell_value is not None:
                    values.append(str(cell_value))
            
            return values
        
        # Handle plain text files (one value per line)
        elif file_path.suffix == '.txt':
            with open(file_path, 'r') as f:
                return [line.strip() for line in f if line.strip()]
        
        else:
            raise ValueError(f"Unsupported file type: {file_path.suffix}")
    
    def __call__(self, func: Callable) -> Callable:
        """Wrap the function to iterate over test values."""
        
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            results = []
            
            # Store original modules/functions that might be patched
            original_funcs = {}
            
            for test_value in self.test_values:
                # Monkey-patch any yaml/config reading in the function's globals
                # This intercepts calls to yaml_open or similar functions
                if hasattr(func, '__globals__'):
                    # Look for common config-reading function patterns
                    for name, obj in func.__globals__.items():
                        if callable(obj) and any(keyword in name.lower() 
                                                for keyword in ['yaml', 'config', 'load', 'read']):
                            if name not in original_funcs:
                                original_funcs[name] = obj
                            
                            # Create a patched version that returns our test value
                            def patched_func(*args, **kwargs):
                                # If the param_name suggests it's a dict key, return a dict
                                if 'dict' in self.param_name.lower() or 'config' in self.param_name.lower():
                                    return {self.param_name: test_value}
                                # Otherwise return the test value directly
                                return test_value
                            
                            func.__globals__[name] = patched_func
                
                # Execute the function
                try:
                    output = func(*args, **kwargs)
                    results.append({
                        self.param_name: test_value,
                        'func_output': output
                    })
                except Exception as e:
                    results.append({
                        self.param_name: test_value,
                        'func_output': f"ERROR: {str(e)}"
                    })
            
            # Restore original functions
            for name, original_func in original_funcs.items():
                func.__globals__[name] = original_func
            
            # Return DataFrame
            return pd.DataFrame(results)
        
        return wrapper


# Alternative approach using context manager for more control
class TestContext:
    """
    Context manager for testing functions with different configurations.
    More explicit control than the decorator approach.
    """
    
    def __init__(self, func_globals: dict, config_func_name: str):
        self.func_globals = func_globals
        self.config_func_name = config_func_name
        self.original_func = None
        
    def __enter__(self):
        self.original_func = self.func_globals.get(self.config_func_name)
        return self
    
    def set_return_value(self, value: Any):
        """Set what the config function should return."""
        self.func_globals[self.config_func_name] = lambda *args, **kwargs: value
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.original_func:
            self.func_globals[self.config_func_name] = self.original_func


# Helper function for manual testing without decorator
def test_function_with_values(func: Callable, param_name: str, 
                              test_values: Union[List[str], str],
                              config_func_name: str = 'yaml_open',
                              **yaml_kwargs) -> pd.DataFrame:
    """
    Test a function with multiple values for a parameter.
    
    Args:
        func: The function to test
        param_name: Name of the parameter being tested
        test_values: List of values or path to file containing values
        config_func_name: Name of the config-reading function to patch
        **yaml_kwargs: Additional kwargs for loading YAML/Excel files
    
    Returns:
        DataFrame with test values and outputs
    """
    decorator = TestIterator(param_name, test_values, **yaml_kwargs)
    wrapped_func = decorator(func)
    return wrapped_func()


# Example usage
if __name__ == "__main__":
    # Mock functions for demonstration
    def yaml_open(filepath):
        return "Hi my name is {name} and i am {age} years old"
    
    def get_age():
        import random
        return random.randint(18, 90)
    
    def some_other_func(template, name):
        age = get_age()
        return template.format(name=name, age=age)
    
    # Example 1: Using decorator with list
    @TestIterator("major_string", [
        "Hi i am {name}, i am {age}",
        "Hey its {name} aged {age}",
        "hello there {age} old {name} this side"
    ])
    def pyfunc(minor_str="Anu"):
        age_value = get_age()
        maj_str = yaml_open("major.yaml")
        final_str = some_other_func(maj_str, minor_str)
        return final_str
    
    # Run the test
    df = pyfunc()
    print(df)
    
    # Example 2: Using with YAML file
    # @TestIterator("major_string", "test_strings.yaml", yaml_key="templates")
    # def pyfunc2(minor_str="Anu"):
    #     ...
    
    # Example 3: Using with Excel file
    # @TestIterator("major_string", "test_strings.xlsx", excel_column="templates")
    # def pyfunc3(minor_str="Anu"):
    #     ...